import openpyxl
from pyautocad import Autocad
from src.classes.points_classes import Row


def get_coordinates(point_group, range_, worksheet):
    for i in range(range_[0], range_[1] + 1):
        coords = []
        for j in range(range_[2], range_[3] + 1, 3):
            coord = [
                worksheet.cell(row=i, column=j).value,
                worksheet.cell(row=i, column=j + 1).value,
                worksheet.cell(row=i, column=j + 2).value,
            ]
            if "None" in coord:
                continue
            coords.append(coord)
        if len(coords) != 0:
            point_group.rows.append(Row(coords))
    return point_group


def launch_autocad(path):
    app = Autocad(create_if_not_exists=True)
    app.Application.Documents.Open(path)
    app.prompt('Hello, Autocad from Python\n')
    print(app.doc.Name)
    return app


def launch_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb


def autocad_save(app, path, worksheet):
    app.ActiveDocument.SaveAs(path +
                              "_" +
                              str(worksheet["W52"].value) +
                              "_" +
                              str(worksheet["U58"].value)
                              )
