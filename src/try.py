import openpyxl
import array
import json
from singleton import*
from pyautocad import Autocad, APoint


class Point:
    def __init__(self, coord):
        self.center = array.array('d', [coord[0], coord[1], coord[2]])
        self.x = coord[0]
        self.y = coord[1]
        self.z = coord[2]

    def offset(self, offset_x, offset_y):
        return array.array('d', [self.x + offset_x, self.y + offset_y, self.z])


class Row:
    def __init__(self, points):
        self.points = list()
        for i in points:
            self.points.append(Point(i))


class Grid:
    def __init__(self):
        self.rows = list()

    def row_len(self, row_num):
        return len(self.rows[row_num].points)

    def grid_len(self):
        result = 0
        for i in range(len(self.rows)):
            result += len(self.rows[i].points)
        return result

    def get_point(self, row, column):
        if row >= len(self.rows):
            raise ValueError('row out of range')
        if column >= len(self.rows[row].points):
            raise ValueError('column out of range')
        return self.rows[row].points[column]

    def get_upper_left(self):
        return self.get_point(0, 0)

    def get_upper_right(self):
        return self.get_point(0, len(self.rows[0].points) - 1)

    def get_down_left(self):
        return self.get_point(len(self.rows) - 1, 0)

    def get_down_right(self):
        last_row = len(self.rows) - 1
        return self.get_point(last_row, len(self.rows[last_row].points) - 1)


class Constructor:
    def __init__(self, app):
        self.app = app

    def add_line(self, point1, point2, layer):
        self.app.model.AddLine(point1, point2).layer = layer

    @staticmethod
    def add_polyline(group, row_num, layer):
        sequence = array.array('d', [group.get_point(row_num, 0).x,
                                     group.get_point(row_num, 0).y,
                                     group.get_point(row_num, 0).z])
        end_point = group.get_point(row_num, group.row_len(row_num) - 1).center
        for j in range(1, group.row_len(row_num)):
            sequence += group.get_point(row_num, j).center
        if len(sequence) <= 6:
            acad.model.AddPolyline(sequence).layer = layer
        else:
            acad.model.AddPolyline(end_point + sequence).layer = layer

    @staticmethod
    def add_block_group(group, block_name):
        for i in range(len(group.rows)):
            for j in range(len(group.rows[i].points)):
                acad.model.InsertBlock(group.get_point(i, j).center, block_name, 1, 1, 1, 0)


class Editor:
    def __init__(self, app):
        self.app = app

    @staticmethod
    def override_dim(group, p1, p2):
        return str(group.row_len(p1[0]) - 1) + \
               "x" + str(int((group.get_point(p2[0], p2[1]).y -
                              group.get_point(p1[0], p1[1]).y) / (group.row_len(p1[0]) - 1))) + \
               "=" + str(group.get_point(p2[0], p2[1]).y - group.get_point(p1[0], p1[1]).y)

    @staticmethod
    def move_front_view():
        acad.ActiveDocument.SendCommand("ПЕРЕНЕСТИ\n10000,-2000\n-1000,2000\n\n0,0\n-12000,0\n")

    @staticmethod
    def move_top_view():
        acad.ActiveDocument.SendCommand("ПЕРЕНЕСТИ\n10000,-2000\n-1000,2000\n\n0,0\n-12000,-5000\n")

    @staticmethod
    def move_side_view():
        acad.ActiveDocument.SendCommand("ПЕРЕНЕСТИ\n10000,-2000\n-1000,2000\n\n0,0\n5000,0\n")

    @staticmethod
    def change_scale(scale):
        return acad.ActiveDocument.SendCommand("CANNOSCALE\n1:" + str(scale) + "\n")


class Dimensions:
    def __init__(self, app):
        self.app = app

    @staticmethod
    def rotated_dim_x(point1, point2, start_point, indent, scale):
        loc = array.array('d', [start_point.x, start_point.y + indent * scale, start_point.z])
        dim = acad.model.AddDimRotated(point1, point2, loc, 0)
        dim.layer = "Размер"
        return dim

    @staticmethod
    def chain_rotated_dim_x(group, start_point, offset, indent, scale):
        for i in range(len(group.rows) - 1):
            Dimensions(acad).rotated_dim_x(group.get_point(i, 0).offset(0, offset),
                                           group.get_point(i + 1, 0).offset(0, offset),
                                           start_point, indent, scale)

    @staticmethod
    def rotated_dim_y(point1, point2, start_point, indent, scale):
        loc = array.array('d', [start_point.x + indent * scale, start_point.y, start_point.z])
        dim = acad.model.AddDimRotated(point1, point2, loc, -1.5708)
        dim.layer = "Размер"
        return dim

    @staticmethod
    def chain_rotated_dim_y(group, start_point, offset, indent, scale):
        for i in range(group.row_len(0) - 1):
            Dimensions(acad).rotated_dim_y(group.get_point(0, i).offset(offset, 0),
                                           group.get_point(0, i + 1).offset(offset, 0),
                                           start_point, indent, scale)


class Design:
    def __init__(self, app, worksheet):
        self.app = app
        self.worksheet = worksheet

    def main_header(self, scale):
        main_header = self.app.model.AddText(u"Конструкция "
                                             + str(self.worksheet["C2"].value) + "a "
                                             + str(self.worksheet["D2"].value)
                                             + " (1:" + str(scale) + ")",
                                             APoint(g['FGP'].get_point(0, 1).x / 2 - 700,
                                                    g['FGP'].get_point(1, 2).y + 600),
                                             4 * scale)
        main_header.layer = "Текст заголовков"
        return main_header

    def header(self, text, group, scale):
        header = self.app.model.AddText(text,
                                        APoint(group.get_point(0, 1).x / 2 - 180,
                                               group.get_point(1, 2).y + 500),
                                        4 * scale)
        header.layer = "Текст заголовков"
        return header

    @staticmethod
    def insert_block(point, block_name, scale):
        return acad.model.InsertBlock(point, block_name, scale, scale, scale, 0)


class Leader:
    def __init__(self, app, worksheet):
        self.app = app
        self.worksheet = worksheet

    def add_mleader_holes(self, point1, point2, group):
        text = "∅" + str(sheet["M2"].value) + "\n" + str(group.grid_len()) + " отв."
        mleader = self.app.model.AddMLeader(point1 + point2, 1)
        mleader.TextString = text
        mleader.TextHeight = 2.5
        mleader.layer = "Размер"
        return mleader

    def add_mleader_pos(self, point1, point2, pos):
        mleader = self.app.model.AddMLeader(point1 + point2, 1)
        mleader.TextString = pos
        mleader.layer = "Размер"
        return mleader


class Tables:
    def __init__(self, app, worksheet):
        self.app = app
        self.worksheet = worksheet

    def add_specification(self):
        table = self.app.find_one(object_name_or_list="table")
        table.SetText(0, 0, "Спецификация " + str(self.worksheet["C2"].value) + "a "
                      + str(self.worksheet["D2"].value))
        table.SetText(18, 0, "* - масса дана с учетом 1,5 % на сварные швы")

        for i in range(32, 40):
            for j in range(18, 22):
                table.SetCellValue(i - 30, j - 18, self.worksheet.cell(row=i, column=j).value)

        for i in range(32, 40, 2):
            table.SetCellValue(i - 30, 5, str(self.worksheet.cell(row=i, column=22).value))
            table.SetCellValue(i - 30, 6, str('%.3f' % self.worksheet.cell(row=i, column=23).value).replace('.', ','))
        table.SetCellValue(2, 7, str('%.3f' % self.worksheet.cell(row=32, column=24).value).replace('.', ','))

        for i in range(2, 18, 2):
            table.MergeCells(i, i + 1, 2, 2)
            if table.GetCellValue(i, 1) == 0:
                table.DeleteRows(i, 17 - i + 1)
                table.MergeCells(2, i - 1, 0, 0)
                table.MergeCells(2, i - 1, 7, 7)
                break


def get_coordinates(point_group, range_):
    for i in range(range_[0], range_[1] + 1):
        coords = []
        for j in range(range_[2], range_[3] + 1, 3):
            coord = [
                sheet.cell(row=i, column=j).value,
                sheet.cell(row=i, column=j + 1).value,
                sheet.cell(row=i, column=j + 2).value,
            ]
            if "None" in coord:
                continue
            coords.append(coord)
        if len(coords) != 0:
            point_group.rows.append(Row(coords))
    return point_group


@singleton
class Settings:
    def __init__(self):
        with open('..\\etc\\config.json', encoding='utf-8') as data_file:
            settings = json.loads(data_file.read())
        points = {name: Grid() for name in settings['groups']}
        for i in settings["groups"]:
            get_coordinates(points[i], settings['groups'][i])

        self.scale_front = settings["scales"]["front"]
        self.scale_top = settings["scales"]["top"]
        self.scale_side = settings["scales"]["side"]
        self.holes_offset = settings["holes_offset"]
        self.points = points


wb = openpyxl.load_workbook('..\\etc\\initial_data.xlsx', data_only=True)
sheet = wb['P-3']

acad = Autocad(create_if_not_exists=True)
acad.Application.Documents.Open('D:\\Element_constructor\\etc\\pattern.dwg')
acad.prompt('Hello, Autocad from Python\n')
print(acad.doc.Name)

initial_data = Settings()
scale_front = initial_data.scale_front
scale_top = initial_data.scale_top
scale_side = initial_data.scale_side
holes_offset = initial_data.holes_offset
g = initial_data.points


def create_front_view(scale):
    Editor(acad).change_scale(scale_front)

    Constructor(acad).add_polyline(g['FGP'], 0, "Основная 0.25")
    Constructor(acad).add_polyline(g['FGP'], 1, "Основная 0.25")
    Constructor(acad).add_polyline(g['FGP'], 2, "Основная 0.25")

    Constructor(acad).add_block_group(g['LWH'], "hole_25")
    Constructor(acad).add_block_group(g['RWH'], "hole_25")

    # bottom line
    Dimensions(acad).rotated_dim_x(g['FGP'].get_point(0, 0).center,
                                   g['LWH'].get_point(0, 0).offset(0, -holes_offset),
                                   g['FGP'].get_point(0, 0), -10, scale)
    Dimensions(acad).chain_rotated_dim_x(g['LWH'], g['FGP'].get_point(0, 0), -holes_offset, -10, scale)
    Dimensions(acad).rotated_dim_x(g['LWH'].get_point(len(g['LWH'].rows)-1, 0).offset(0, -holes_offset),
                                   g['RWH'].get_point(len(g['RWH'].rows)-1, 0).offset(0, -holes_offset),
                                   g['FGP'].get_point(0, 0), -10, scale)
    Dimensions(acad).chain_rotated_dim_x(g['RWH'], g['FGP'].get_point(0, 0), -holes_offset, -10, scale)
    Dimensions(acad).rotated_dim_x(g['RWH'].get_point(0, 0).offset(0, -holes_offset),
                                   g['FGP'].get_point(0, 1).center,
                                   g['FGP'].get_point(0, 0), -10, scale)

    # left line
    Dimensions(acad).rotated_dim_y(g['FGP'].get_point(0, 0).center,
                                   g['LWH'].get_point(0, 0).offset(-holes_offset, 0),
                                   g['FGP'].get_point(0, 0), -10, scale)
    # ======================================
    Dimensions(acad).rotated_dim_y(g['LWH'].get_point(0, 0).offset(-holes_offset, 0),
                                   g['LWH'].get_upper_right().offset(-holes_offset, 0),
                                   g['FGP'].get_point(0, 0), -10, scale).\
        TextOverride = Editor(acad).override_dim(g['LWH'], [0, 0], [0, g['LWH'].row_len(0)-1])
    # ======================================
    Dimensions(acad).rotated_dim_y(g['FGP'].get_point(1, 3).center,
                                   g['LWH'].get_upper_right().offset(-holes_offset, 0),
                                   g['FGP'].get_point(0, 0), -10, scale)

    # right line
    Dimensions(acad).rotated_dim_y(g['FGP'].get_point(0, 1).center,
                                   g['RWH'].get_point(0, 0).offset(holes_offset, 0),
                                   g['FGP'].get_point(0, 1), 10, scale)
    Dimensions(acad).rotated_dim_y(g['RWH'].get_point(0, 0).offset(holes_offset, 0),
                                   g['RWH'].get_upper_right().offset(holes_offset, 0),
                                   g['FGP'].get_point(0, 1), 10, scale).\
        TextOverride = Editor(acad).override_dim(g['LWH'], [0, 0], [0, g['LWH'].row_len(0)-1])

    Dimensions(acad).rotated_dim_y(g['FGP'].get_point(1, 2).center,
                                   g['RWH'].get_upper_right().offset(holes_offset, 0),
                                   g['FGP'].get_point(0, 1), 10, scale)

    # top line
    Dimensions(acad).rotated_dim_x(g['FGP'].get_point(1, 3).center,
                                   g['FGP'].get_point(1, 2).center,
                                   g['FGP'].get_point(1, 2), 10, scale)

    front_header = Design(acad, sheet)
    front_header.main_header(scale_front)

    holes1 = Leader(acad, sheet, )
    holes1.add_mleader_holes(g['LWH'].get_down_left().center,
                             g['FGP'].get_upper_left().
                             offset(g['LWH'].get_down_left().x + 5 * scale_front, -5 * scale_front),
                             g['LWH'])
    holes2 = Leader(acad, sheet)
    holes2.add_mleader_holes(g['RWH'].get_upper_right().center,
                             g['FGP'].get_point(1, 2).
                             offset(5 * scale_front, 10 * scale_front),
                             g['RWH'])
    pos1 = Leader(acad, sheet)
    pos1.add_mleader_pos(g['FGP'].get_point(1, 3).offset(1000, 0),
                         g['FGP'].get_point(1, 3).offset(1000, 13 * scale_front), "2")
    if g['TGP'].get_point(0, 1).center != g['TGP'].get_point(1, 1).center:
        text_pos1 = "3"
    else:
        text_pos1 = "2"
    pos2 = Leader(acad, sheet)
    pos2.add_mleader_pos(g['FGP'].get_upper_left().offset(1000, 0),
                         g['FGP'].get_upper_left().offset(1000, -17 * scale_front), text_pos1)
    pos3 = Leader(acad, sheet)
    pos3.add_mleader_pos(g['FGP'].get_point(0, 1).offset(0, 80),
                         g['FGP'].get_point(0, 1).offset(5 * scale_front, -10 * scale_front), "1")
    Design(acad, sheet).insert_block(g['FGP'].get_point(1, 3).offset(5, 15 * scale_front),
                                     "sec_1_t", scale_front)
    Design(acad, sheet).insert_block(g['FGP'].get_upper_left().offset(5, -15 * scale_front),
                                     "sec_1_b", scale_front)
    Design(acad, sheet).insert_block(g['FGP'].get_point(1, 2).
                                     offset(float(-g['FGP'].get_point(1, 2).x / 2), 15 * scale_front),
                                     "view_A_t", scale_front)
    Design(acad, sheet).insert_block(g['FGP'].get_point(0, 1).
                                     offset(float(-g['FGP'].get_point(0, 1).x / 2), -15 * scale_front),
                                     "view_po_A_b", scale_front)
    Editor(acad).move_front_view()


def create_top_view(scale):
    Editor(acad).change_scale(scale_top)
    Constructor(acad).add_block_group(g['LCH'], "hole_25")
    Constructor(acad).add_block_group(g['MCH'], "hole_25")
    Constructor(acad).add_block_group(g['RCH'], "hole_25")
    Constructor(acad).add_polyline(g['TGP'], 0, "Основная 0.25")
    if g['TGP'].get_point(0, 1).center != g['TGP'].get_point(1, 1).center:
        Constructor(acad).add_polyline(g['TGP'], 1, "Основная 0.25")
    Constructor(acad).add_polyline(g['TGP'], 2, "Ось")
    Constructor(acad).add_polyline(g['TGP'], 3, "Пунктир")

    # bottom line
    Dimensions(acad).rotated_dim_x(g['TGP'].get_point(0, 0).center,
                                   g['LCH'].get_point(0, 0).offset(0, -holes_offset),
                                   g['TGP'].get_point(0, 0), -10, scale)
    Dimensions(acad).chain_rotated_dim_x(g['LCH'], g['TGP'].get_point(0, 0), -holes_offset, -10, scale)
    if len(g['MCH'].rows) > 0:
        Dimensions(acad).rotated_dim_x(g['LCH'].get_point(len(g['LCH'].rows) - 1, 0).offset(0, -holes_offset),
                                       g['MCH'].get_point(0, 0).offset(0, -holes_offset),
                                       g['TGP'].get_point(0, 0), -10, scale)

        Dimensions(acad).chain_rotated_dim_x(g['MCH'], g['TGP'].get_point(0, 0), -holes_offset, -10, scale)
        Dimensions(acad).rotated_dim_x(g['MCH'].get_point(len(g['MCH'].rows)-1, 0).offset(0, -holes_offset),
                                       g['RCH'].get_point(len(g['RCH'].rows)-1, 0).offset(0, -holes_offset),
                                       g['TGP'].get_point(0, 0), -10, scale)
    else:
        Dimensions(acad).rotated_dim_x(g['LCH'].get_point(len(g['LCH'].rows) - 1, 0).offset(0, -holes_offset),
                                       g['RCH'].get_point(len(g['RCH'].rows) - 1, 0).offset(0, -holes_offset),
                                       g['TGP'].get_point(0, 0), -10, scale)
    Dimensions(acad).chain_rotated_dim_x(g['RCH'], g['TGP'].get_point(0, 0), -holes_offset, -10, scale)
    Dimensions(acad).rotated_dim_x(g['RCH'].get_point(0, 0).offset(0, -holes_offset),
                                   g['TGP'].get_point(0, 1).center,
                                   g['TGP'].get_point(0, 0), -10, scale)

    # left line
    Dimensions(acad).rotated_dim_y(g['TGP'].get_point(0, 0).center,
                                   g['LCH'].get_point(0, 0).offset(-holes_offset, 0),
                                   g['TGP'].get_point(0, 0), -10, scale)
    Dimensions(acad).chain_rotated_dim_y(g['LCH'], g['TGP'].get_point(0, 0), -holes_offset, -10, scale)
    Dimensions(acad).rotated_dim_y(g['TGP'].get_upper_right().center,
                                   g['LCH'].get_upper_right().offset(-holes_offset, 0),
                                   g['TGP'].get_point(0, 0), -10, scale)

    # right line
    Dimensions(acad).rotated_dim_y(g['TGP'].get_point(0, 1).center,
                                   g['RCH'].get_point(0, 0).offset(holes_offset, 0),
                                   g['TGP'].get_point(0, 1), 10, scale)
    Dimensions(acad).chain_rotated_dim_y(g['RCH'], g['TGP'].get_point(0, 1), holes_offset, 10, scale)
    Dimensions(acad).rotated_dim_y(g['TGP'].get_point(0, 2).center,
                                   g['RCH'].get_upper_right().offset(holes_offset, 0),
                                   g['TGP'].get_point(0, 1), 10, scale)

    # top line
    Dimensions(acad).rotated_dim_x(g['TGP'].get_point(0, 3).center,
                                   g['TGP'].get_point(0, 2).center,
                                   g['TGP'].get_point(0, 2), 10, scale)

    # center line
    if g['MCH'].grid_len() > 0:
        Dimensions(acad).rotated_dim_y(g['TGP'].get_point(0, 0).
                                       offset(g['MCH'].get_upper_left().x - 200, 0),
                                       g['MCH'].get_point(0, 0).offset(-holes_offset, 0),
                                       g['MCH'].get_point(0, 0), -10, scale)
        Dimensions(acad).chain_rotated_dim_y(g['MCH'], g['MCH'].get_point(0, 0), -holes_offset, -10, scale)
        Dimensions(acad).rotated_dim_y(g['TGP'].get_upper_right().
                                       offset(g['MCH'].get_upper_left().x - 200, 0),
                                       g['MCH'].get_upper_right().offset(-holes_offset, 0),
                                       g['MCH'].get_point(0, 0), -10, scale)
        holes2 = Leader(acad, sheet)
        holes2.add_mleader_holes(g['MCH'].get_down_right().center,
                                 g['TGP'].get_point(0, 3).
                                 offset(g['MCH'].get_down_right().x + 5 * scale_top, 5 * scale_top), g['MCH'])

    top_header = Design(acad, sheet)
    top_header.header(u"A (1:" + str(scale_top) + ")", g['TGP'], scale_top)
    holes1 = Leader(acad, sheet)
    holes1.add_mleader_holes(g['LCH'].get_point(len(g['LCH'].rows) - 1, 0).center,
                             g['TGP'].get_point(0, 0).
                             offset(g['LCH'].get_point(len(g['LCH'].rows) - 1, 0).x +
                                    5 * scale_top, - 5 * scale_top), g['LCH'])
    holes3 = Leader(acad, sheet)
    holes3.add_mleader_holes(g['RCH'].get_upper_right().center,
                             g['TGP'].get_point(0, 2).
                             offset(5 * scale_top, 10 * scale_top), g['RCH'])
    if g['TGP'].get_point(0, 1).center != g['TGP'].get_point(1, 1).center:
        text_pos2 = "2(3)"
    else:
        text_pos2 = "2"
    pos1 = Leader(acad, sheet)
    pos1.add_mleader_pos(g['TGP'].get_upper_left().offset(1000, 0),
                         g['TGP'].get_upper_left().offset(1000, -17 * scale_front), text_pos2)
    Editor(acad).move_top_view()


def create_side_view(scale):
    Editor(acad).change_scale(scale_side)

    Constructor(acad).add_polyline(g['SGP'], 0, "Основная 0.25")
    Constructor(acad).add_polyline(g['SGP'], 1, "Основная 0.25")
    Constructor(acad).add_polyline(g['SGP'], 2, "Ось")
    Constructor(acad).add_polyline(g['SGP'], 3, "Основная 0.25")

    # bottom line
    if g['SGP'].get_point(0, 1).x != g['SGP'].get_point(1, 1).x:
        Dimensions(acad).rotated_dim_x(g['SGP'].get_point(0, 0).center,
                                       g['SGP'].get_point(0, 1).center,
                                       g['SGP'].get_point(0, 0), -10, scale)

    # left line
    Dimensions(acad).rotated_dim_y(g['SGP'].get_point(0, 0).center,
                                   g['SGP'].get_point(0, g['SGP'].row_len(0) - 1).center,
                                   g['SGP'].get_point(0, 0), -10, scale)
    Dimensions(acad).rotated_dim_y(g['SGP'].get_point(0, g['SGP'].row_len(0) - 1).center,
                                   g['SGP'].get_point(1, 0).center,
                                   g['SGP'].get_point(0, 0), -10, scale)
    Dimensions(acad).rotated_dim_y(g['SGP'].get_point(1, 0).center,
                                   g['SGP'].get_point(1, g['SGP'].row_len(1) - 1).center,
                                   g['SGP'].get_point(0, 0), -10, scale)

    # right line
    Dimensions(acad).rotated_dim_y(g['SGP'].get_point(0, 1).center,
                                   g['SGP'].get_point(1, 2).center,
                                   g['SGP'].get_point(0, 1), 10, scale)

    # top line
    Dimensions(acad).rotated_dim_x(g['SGP'].get_point(1, 2).center,
                                   g['SGP'].get_point(1, g['SGP'].row_len(0) - 1).center,
                                   g['SGP'].get_point(1, 2), 10, scale)

    side_header = Design(acad, sheet)
    side_header.header(u"1-1 (1:" + str(scale_side) + ")", g['SGP'], scale_side)
    pos1 = Leader(acad, sheet)
    pos1.add_mleader_pos(g['SGP'].get_point(1, 2).offset(-2 * scale_side, 0),
                         g['SGP'].get_point(1, 2).offset(7 * scale_side, 10 * scale_side), "2")
    if g['TGP'].get_point(0, 1).center != g['TGP'].get_point(1, 1).center:
        text_pos3 = "3"
    else:
        text_pos3 = "2"
    pos2 = Leader(acad, sheet)
    pos2.add_mleader_pos(g['SGP'].get_point(0, 1).offset(-2 * scale_side, 0),
                         g['SGP'].get_point(0, 1).offset(7 * scale_side, -10 * scale_side), text_pos3)
    pos3 = Leader(acad, sheet)
    pos3.add_mleader_pos(g['SGP'].get_point(3, 2).offset(0, -g['SGP'].get_point(3, 2). y / 2),
                         g['SGP'].get_point(3, 2).offset(7 * scale_side,
                                                         -10 * scale_side
                                                         - g['SGP'].get_point(3, 2). y / 2), "1")

    Editor(acad).move_side_view()


create_front_view(scale_front)
create_top_view(scale_top)
create_side_view(scale_side)

table1 = Tables(acad, sheet)
table1.add_specification()
