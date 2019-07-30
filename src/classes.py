import openpyxl
import array
from pyautocad import Autocad, APoint, ACAD


class Tables:
    def __init__(self, app, worksheet):
        self.app = app
        self.worksheet = worksheet

    def add_specification(self):
        table = self.app.find_one(object_name_or_list="table")
        table.SetText(0, 0, "Спецификация " + str(self.worksheet["C2"].value) + "и "
                      + str(self.worksheet["D2"].value))
        table.SetText(18, 0, "* - масса дана с учетом 2% на сварные швы")

        for i in range(32, 40):
            for j in range(18, 22):
                table.SetCellValue(i - 30, j - 18, self.worksheet.cell(row=i, column=j).value)

        for i in range(32, 40, 2):
            table.SetCellValue(i - 30, 5, str(self.worksheet.cell(row=i, column=22).value))
            table.SetCellValue(i - 30, 6, str('%.3f' % self.worksheet.cell(row=i, column=23).value).replace('.', ','))
        table.SetCellValue(2, 7, str('%.3f' % self.worksheet.cell(row=32, column=24).value).replace('.', ','))

        for i in range(2, 18, 2):
            table.MergeCells(i, i + 1, 2, 2)
            if table.GetCellValue(i, 1) == 0:
                table.DeleteRows(i, 17 - i + 1)
                table.MergeCells(2, i - 1, 0, 0)
                table.MergeCells(2, i - 1, 7, 7)
                break

    def add_stamp(self):
        self.app.ActiveDocument.SendCommand('РЛИСТ\nТ\nГотовое\n')
        table = self.app.find_one(object_name_or_list="table")
        table.SetText(8, 6, str(self.worksheet["U58"].value))
        table.SetText(6, 8, str(self.worksheet["W55"].value))
        text = self.app.doc.PaperSpace.AddText(self.worksheet["W52"].value, APoint(0, 0), 2.5)
        text.Alignment = ACAD.acAlignmentMiddleCenter
        text.TextAlignmentPoint = APoint(520, 229.50)


def get_coordinates(point_group, range_, worksheet):
    for i in range(range_[0], range_[1] + 1):
        coords = []
        for j in range(range_[2], range_[3] + 1, 3):
            coord = [
                worksheet.cell(row=i, column=j).value,
                worksheet.cell(row=i, column=j + 1).value,
                worksheet.cell(row=i, column=j + 2).value,
            ]
            if "None" in coord:
                continue
            coords.append(coord)
        if len(coords) != 0:
            point_group.rows.append(Row(coords))
    return point_group


def launch_autocad(path):
    app = Autocad(create_if_not_exists=True)
    app.Application.Documents.Open(path)
    app.prompt('Hello, Autocad from Python\n')
    print(app.doc.Name)
    return app


def launch_excel(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb


def autocad_save(app, path, worksheet):
    app.ActiveDocument.SaveAs(path +
                              "_" +
                              str(worksheet["W52"].value) +
                              "_" +
                              str(worksheet["U58"].value)
                              )
