from src.classes import i_beam_front_view, i_beam_top_view, i_beam_side_view, holes_grid
from pyautocad import Autocad
import openpyxl


print("""Welcome to the \"Bridge_truss\" module !!!`
You can draw some blueprints using this module.
You can get initial data from MS Excel or type 
necessary parameters right now.
""")

wb = openpyxl.load_workbook("D:\\Element_constructor\\etc\\initial_data.xlsx", data_only=True)
answer = input("Do you want to open all the sheet from the initial data? Y/N ")
if answer.lower() == "y":
    for i in range(len(wb.sheetnames)):
        sheet = wb[wb.sheetnames[i]]
        choice_excel = sheet['C2'].value
        if choice_excel == 1:

            print("You chose the I-beam element. Processing initial data...")

            acad = Autocad(create_if_not_exists=True)
            acad.Application.Documents.Open("D:\\Element_constructor\\etc\\pattern.dwg")
            acad.prompt("Hello, Autocad from Python\n")
            print(acad.doc.Name)

            scale_front = 20
            scale_top = 20
            scale_side = 10
            offset_view_y = -5000
            offset_view_x = 15000
            l_i = sheet['D2'].value
            b1_i = sheet['E2'].value
            tw1_i = sheet['F2'].value
            b2_i = sheet['G2'].value
            tw2_i = sheet['H2'].value
            h_i = sheet['I2'].value
            t_i = sheet['J2'].value

            i_beam_front_view(l_i, tw1_i, tw2_i, h_i, sheet, scale_front)
            i_beam_top_view(l_i, b1_i, tw1_i, b2_i, tw2_i, h_i, t_i, sheet, scale_top, offset_view_y)
            i_beam_side_view(l_i, b1_i, tw1_i, b2_i, tw2_i, h_i, t_i, sheet, scale_side, offset_view_x)
            holes_grid(l_i, b1_i, tw1_i, b2_i, tw2_i, h_i, t_i, sheet, scale_front, scale_top, scale_side,
                       offset_view_y, offset_view_x)
else:
    sheet_name = input("Enter a name of a sheet: ")
    sheet = wb[sheet_name]
    choice_excel = sheet['C2'].value
    if choice_excel == 1:
        print("You chose the I-beam element. Processing initial data...")

        acad = Autocad(create_if_not_exists=True)
        acad.Application.Documents.Open("D:\\Element_constructor\\etc\\pattern.dwg")
        acad.prompt("Hello, Autocad from Python\n")
        print(acad.doc.Name)

        scale_front = 20
        scale_top = 20
        scale_side = 10
        offset_view_y = -5000
        offset_view_x = 15000
        l_i = sheet['D2'].value
        b1_i = sheet['E2'].value
        tw1_i = sheet['F2'].value
        b2_i = sheet['G2'].value
        tw2_i = sheet['H2'].value
        h_i = sheet['I2'].value
        t_i = sheet['J2'].value

        i_beam_front_view(l_i, tw1_i, tw2_i, h_i, sheet, scale_front)
        i_beam_top_view(l_i, b1_i, tw1_i, b2_i, tw2_i, h_i, t_i, sheet, scale_top, offset_view_y)
        i_beam_side_view(l_i, b1_i, tw1_i, b2_i, tw2_i, h_i, t_i, sheet, scale_side, offset_view_x)
        holes_grid(l_i, b1_i, tw1_i, b2_i, tw2_i, h_i, t_i, sheet, scale_front, scale_top, scale_side,
                   offset_view_y, offset_view_x)

print("Keep working ...")
