from pyautocad import Autocad, APoint, types
import comtypes
import array
import openpyxl


acad = Autocad(create_if_not_exists=True)
# acad.Application.Documents.Open("D:\\IT\\Bridge_truss\\pattern.dwg")
acad.prompt("Hello, Autocad from Python\n")
print(acad.doc.Name)


def dim_aligned(point1, point2, scale, indent_x=0, indent_y=0, start_point_x=0, start_point_y=0, move_x=0, move_y=0):
    """ This function places an aligned dimension.
        Parameters:
        1. point1, point2 - the points between which you want to place a dimension.
        2. scale - a scale you want to create the current dimension.
        3. indent_x, indent_y - indents in accordance with GOST:
          - 10 mm for first step;
          - 8 mm for other steps.
          For 2-nd line of dimensions indent = 10+8, for 3-rd -10+8+8, etc.

          !!! BE CAREFUL !!! with "+" and "-" !!!

        4. start_point_x, start_point_y  - x and y-coordinates from which you need to make an indent.
        5. move_x, move_y - view's offset.
    """
    # It's measured from 0, 0, 0 !!!
    loc = array.array('d', [start_point_x + indent_x * scale, start_point_y + indent_y * scale, 0])
    # ==============================
    dim = acad.model.AddDimAligned(point1, point2, loc)
    dim.layer = "Размер"
    if move_x != 0 or move_y != 0:
        dim.move(APoint(0, 0), APoint(move_x, move_y))


def holes(diameter: str, start_point_x, start_point_y, holes_grid, move_x=0, move_y=0):

    for x in range(0, len(holes_grid), 2):
        block = acad.model.InsertBlock(APoint(start_point_x + holes_grid[x],
                                                  start_point_y + holes_grid[x+1]),
                                           "hole_" + str(diameter), 1, 1, 1, 0)
        block.move(APoint(0, 0), APoint(move_x, move_y))


def i_beam(l_i, b1_i, tw1_i, b2_i, tw2_i, h_i, t_i, sheet):
    """ This function creates the geometry
        of the I-beam element in Autocad
    """

    # Creating front view.

    scale_front = int(input("Enter a scale of front view: "))

    print("Creating of front view...")

    p1 = array.array('d', [0, 0, 0])
    p2 = array.array('d', [l_i, 0, 0])
    p3 = array.array('d', [l_i, tw2_i, 0])
    p4 = array.array('d', [0, tw2_i, 0])
    p5 = array.array('d', [0, h_i - tw1_i, 0])
    p6 = array.array('d', [l_i, h_i - tw1_i, 0])
    p7 = array.array('d', [l_i, h_i, 0])
    p8 = array.array('d', [0, h_i, 0])
    poly_f_b1 = acad.model.AddPolyline(p1 + p2 + p3 + p4 + p1)
    poly_f_b2 = acad.model.AddPolyline(p5 + p6 + p7 + p8 + p5)
    poly_f_h1 = acad.model.AddPolyline(p3 + p4 + p5 + p6 + p3)
    poly_f_b1.layer = "Основная 0.25"
    poly_f_b2.layer = "Основная 0.25"
    poly_f_h1.layer = "Основная 0.25"
    dim_aligned(p1, p2, scale_front, indent_y=-10)
    dim_aligned(p2, p3, scale_front, indent_x=10, start_point_x=l_i)

    # Creating top view

    scale_top = int(input("Enter a scale of the top view: "))
    if scale_front != scale_top:
        input("Please change a scale in model and press \"Y\": ")
    offset_view_y = -5000

    print("Creating of top view...")

    p9 = array.array('d', [0, b2_i, 0])
    p10 = array.array('d', [l_i, b2_i, 0])
    p11 = array.array('d', [-1*scale_top, b2_i / 2, 0])
    p12 = array.array('d', [l_i + 1*scale_top, b2_i / 2, 0])
    p13 = array.array('d', [0, b2_i / 2 - t_i/2, 0])
    p14 = array.array('d', [l_i, b2_i / 2 - t_i/2, 0])
    p15 = array.array('d', [0, b2_i / 2 + t_i / 2, 0])
    p16 = array.array('d', [l_i, b2_i / 2 + t_i / 2, 0])
    poly_t_1 = acad.model.AddPolyline(p1 + p2 + p10 + p9 + p1)
    poly_t_2 = acad.model.AddPolyline(p11 + p12)
    poly_t_3 = acad.model.AddPolyline(p13 + p14)
    poly_t_4 = acad.model.AddPolyline(p15 + p16)
    poly_t_1.layer = "Основная 0.25"
    poly_t_2.layer = "Ось"
    poly_t_3.layer = "Пунктир"
    poly_t_4.layer = "Пунктир"
    poly_t_1.move(APoint(0, 0), APoint(0, offset_view_y))
    poly_t_2.move(APoint(0, 0), APoint(0, offset_view_y))
    poly_t_3.move(APoint(0, 0), APoint(0, offset_view_y))
    poly_t_4.move(APoint(0, 0), APoint(0, offset_view_y))

    print("Creating of bolt holes grid ...")

    # Making holes grid for chords:

    # Bolt holes diameter
    diameter = sheet["K2"].value

    # Left grid

    holes_left = []
    for i in range(18, 24):
        for j in range(2, 16):
            if sheet.cell(row=i, column=j).value is not None:
                holes_left.append(sheet.cell(row=i, column=j).value)

    # Right grid

    holes_right = [0] * len(holes_left)
    for i in range(0, len(holes_left), 2):
        holes_right[i] = -holes_left[i]
    for i in range(1, len(holes_left), 2):
        holes_right[i] = holes_left[i]

    # Middle grid

    holes_mid = []
    for i in range(39, 45):
        for j in range(2, 16):
            if sheet.cell(row=i, column=j).value is not None:
                holes_mid.append(sheet.cell(row=i, column=j).value)
    start_point_x = sheet["B37"].value

    # Making holes grid for wall

    # Left grid

    holes_wall_left = []
    for i in range(60, 66):
        for j in range(2, 16):
            if sheet.cell(row=i, column=j).value is not None:
                holes_wall_left.append(sheet.cell(row=i, column=j).value)

    # Right grid

    holes_wall_right = [0] * len(holes_wall_left)
    for i in range(0, len(holes_wall_left), 2):
        holes_wall_right[i] = -holes_wall_left[i]
    for i in range(1, len(holes_wall_left), 2):
        holes_wall_right[i] = holes_wall_left[i]

    # Creating of bolt holes grid
    holes(diameter, 0, 0, holes_left, move_y=offset_view_y)
    holes(diameter, l_i, 0, holes_right, move_y=offset_view_y)
    holes(diameter, 0, tw2_i, holes_wall_left)
    holes(diameter, l_i, tw2_i, holes_wall_right)
    holes(diameter, start_point_x, 0, holes_mid, move_y=offset_view_y)
    # holes("25", 0, 0, holes_grid3, move_y=offset_view_y)
    # holes("25", 0, 0, holes_grid4, move_y=offset_view_y)
    dim_aligned(p1, p2, scale_top, indent_y=-10, move_y=offset_view_y)

    # Bolt holes side view

    h_s_x = []
    # holes_side front view
    for i in range(68, 88):
        if sheet.cell(row=i, column=2).value is not None:
            h_s_x.append(sheet.cell(row=i, column=2).value)
    for i in range(len(h_s_x)-1):
        p_x1 = array.array('d', [h_s_x[i], 0, 0])
        if h_s_x[i] > 0:
            hole_center1 = acad.model.AddLine(APoint(h_s_x[i], -1 * scale_front),
                                              APoint(h_s_x[i], tw2_i + 1 * scale_front))
            hole_line1 = acad.model.AddLine(APoint(float(h_s_x[i]-diameter/2), 0),
                                            APoint(float(h_s_x[i]-diameter/2), tw2_i))
            hole_line2 = acad.model.AddLine(APoint(float(h_s_x[i]+diameter/2), 0),
                                            APoint(float(h_s_x[i]+diameter/2), tw2_i))
            hole_center1.layer = "Ось"
            hole_line1.layer = "Пунктир"
            hole_line2.layer = "Пунктир"
            hole_center1.copy().move(APoint(0, 0), APoint(0, h_i - tw1_i))
            hole_line1.copy().move(APoint(0, 0), APoint(0, h_i - tw1_i))
            hole_line2.copy().move(APoint(0, 0), APoint(0, h_i - tw1_i))
        p_x2 = array.array('d', [h_s_x[i+1], 0, 0])
        dim_aligned(p_x1, p_x2, scale_front, indent_y=-10, move_y=offset_view_y)

    h_s_s = []
    # holes_side side view

    for i in range(18, 22, 2):
        h_s_s = [].append(sheet.cell(row=i, column=2).value)
# acad.ActiveDocument.SaveAs('D:\\IT\\Bridge_truss\\1.dwg')
